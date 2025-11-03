from openpyxl import*
from tkinter import*

#variaveis
windows = Tk()

def save_profession():
    global name, age, profession, ws, button, entry3
    
    profession = entry3.get()
    print(profession)

    button.config(text="Informações salvas com sucesso!", state=DISABLED)

    #inserir informações
    ws['A2'] = name
    ws['B2'] = age
    ws['C2'] = profession
    wb.save("informações.xlsx")

def save_age():
    global age, entry3
    age = entry2.get()
    print(age)

    button.config(command=save_profession)
    entry3 = Entry(windows, font=("Ink Free",20))
    entry3.pack()

def save_name():
    global name, entry2
    name = entry1.get()
    print(name)

    button.config(command=save_age)
    entry2 = Entry(windows, font=("Ink Free",20))
    entry2.pack()


entry1 = Entry(windows, font=("Ink Free",20))

button = Button(windows, text="Salvar informações", font=("Arial",20), command=save_name)
button.pack()

wb = Workbook()
ws = wb.active

ws['A1'] = "Nome"
ws['B1'] = "Idade"
ws['C1'] = "profissão"

#salvar arquivo e aparecer
entry1.pack()
windows.mainloop()